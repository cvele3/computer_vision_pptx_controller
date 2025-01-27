import csv
import uuid
import openpyxl
import os
import psutil
import time
import win32com.client
import pyautogui
import random

import demo_workflow_specific as specific_flow
import demo_workflows_counting as counting_flow

EXCEL_FILE = "logs/user_performance.xlsx"
ERROR_LOG_FILE = "logs/error_log.csv"

def close_pptx():
    """
    Close the PowerPoint application by terminating its process.
    """
    ppt_app = win32com.client.Dispatch("PowerPoint.Application")
    ppt_app.Quit()

def save_to_excel(user_uuid, results):
    """
    Save the results to the Excel file.
    """
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    # Find the next available row
    row = ws.max_row + 1

    # Write the results to the Excel file
    ws.cell(row=row, column=1, value=str(user_uuid))
    for i, (elapsed_time, error_count) in enumerate(results):
        ws.cell(row=row, column=2 + i * 2, value=elapsed_time)
        ws.cell(row=row, column=3 + i * 2, value=error_count)

    wb.save(EXCEL_FILE)

def main():
    random_uuid = uuid.uuid4()
    specific_results = []
    counting_results = []

    # Initialize error log
    with open(ERROR_LOG_FILE, mode='w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(["Timestamp", "Gesture type", "Flow number", "Step", "Detected Gesture", "Expected Gesture"])

    # Randomize the order of the loops
    if random.choice([True, False]):
        # Process workflows from `specific_flow.WORKFLOWS` first
        for i, workflow in enumerate(specific_flow.WORKFLOWS):
            file_number = i + 1
            pptx_path = f"D:/Ostalo/faks/4. semestar/HCI/computer_vision_pptx_controller/pptx/WORKFLOW_{file_number}.pptx"
            print(f"Processing workflow {i + 1}/{len(specific_flow.WORKFLOWS)}: {' -> '.join(workflow)}")

            # Open the PowerPoint file
            os.startfile(pptx_path)

            time.sleep(5)

            pyautogui.click()

            try:
                # Process the workflow
                elapsed_time, error_count = specific_flow.process_workflow(workflow, i + 1)
                specific_results.append((elapsed_time, error_count))
            except Exception as e:
                print(f"Error processing workflow {i + 1}: {e}")
            finally:
                # Ensure PowerPoint is fully closed before proceeding to the next iteration
                close_pptx()

        # Process workflows from `counting_flow.WORKFLOWS`
        for i, workflow in enumerate(counting_flow.WORKFLOWS):
            file_number = i + 1
            pptx_path = f"D:/Ostalo/faks/4. semestar/HCI/computer_vision_pptx_controller/pptx/WORKFLOW_{file_number}.pptx"
            print(f"Processing workflow {i + 1}/{len(counting_flow.WORKFLOWS)}: {' -> '.join(workflow)}")

            # Open the PowerPoint file
            os.startfile(pptx_path)

            time.sleep(5)

            pyautogui.click()

            try:
                # Process the workflow
                elapsed_time, error_count = counting_flow.process_workflow(workflow, i + 1)
                counting_results.append((elapsed_time, error_count))
            except Exception as e:
                print(f"Error processing workflow {i + 1}: {e}")
            finally:
                # Ensure PowerPoint is fully closed before proceeding to the next iteration
                close_pptx()
    else:
        # Process workflows from `counting_flow.WORKFLOWS` first
        for i, workflow in enumerate(counting_flow.WORKFLOWS):
            file_number = i + 1
            pptx_path = f"D:/Ostalo/faks/4. semestar/HCI/computer_vision_pptx_controller/pptx/WORKFLOW_{file_number}.pptx"
            print(f"Processing workflow {i + 1}/{len(counting_flow.WORKFLOWS)}: {' -> '.join(workflow)}")

            # Open the PowerPoint file
            os.startfile(pptx_path)

            time.sleep(5)

            pyautogui.click()

            try:
                # Process the workflow
                elapsed_time, error_count = counting_flow.process_workflow(workflow, i + 1)
                counting_results.append((elapsed_time, error_count))
            except Exception as e:
                print(f"Error processing workflow {i + 1}: {e}")
            finally:
                # Ensure PowerPoint is fully closed before proceeding to the next iteration
                close_pptx()

        # Process workflows from `specific_flow.WORKFLOWS`
        for i, workflow in enumerate(specific_flow.WORKFLOWS):
            file_number = i + 1
            pptx_path = f"D:/Ostalo/faks/4. semestar/HCI/computer_vision_pptx_controller/pptx/WORKFLOW_{file_number}.pptx"
            print(f"Processing workflow {i + 1}/{len(specific_flow.WORKFLOWS)}: {' -> '.join(workflow)}")

            # Open the PowerPoint file
            os.startfile(pptx_path)

            time.sleep(5)

            pyautogui.click()

            try:
                # Process the workflow
                elapsed_time, error_count = specific_flow.process_workflow(workflow, i + 1)
                specific_results.append((elapsed_time, error_count))
            except Exception as e:
                print(f"Error processing workflow {i + 1}: {e}")
            finally:
                # Ensure PowerPoint is fully closed before proceeding to the next iteration
                close_pptx()

    # Save results to Excel
    save_to_excel(random_uuid, specific_results + counting_results)


if __name__ == "__main__":
    main()